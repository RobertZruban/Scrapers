

import openpyxl  
import pprint
pp = pprint.PrettyPrinter(indent=4)
import random
import requests
import urllib.request
import time
from bs4 import BeautifulSoup
import folium

from selenium import webdriver  
from selenium.common.exceptions import NoSuchElementException  
from selenium.webdriver.common.keys import Keys  
from bs4 import BeautifulSoup
import pandas as pd
df = pd.DataFrame()
url = 'https://www.reality.sk/bratislavsky-kraj/virtualne-prehliadky/?page='
main_url = 'https://www.reality.sk'
#'https://mobilne-telefony.heureka.sk', 'https://inteligentne-hodinky.heureka.sk/', 'https://televizor.heureka.sk/', 'https://sluchadla.heureka.sk/'
               


   
   
latitude_direction = []
longitude_direction = []
cena = []
Druh = []
Typ = []
Plocha = []
nazov = []
text = []
Nazov = []
Celková_podlahová_plocha = []
Nadzemné_podlažie = []
Stav_nehnuteľnosti = []
Forma_vlastníctva = []
Konštrukcia_bytu = []
Rok_výstavby = []
Terasa = []
Energetický_certifikát_budovy = []
Internet = []
Káblová_televízia = []
ALL_data = {}  



 




browser = webdriver.Chrome(r'C:\Users\roboz.DESKTOP-F86F289\Desktop\chromedrive\chromedriver.exe')
links = []
 
count  = 0
pokus = []  

for x in range(1,300):
     
    browser.get(url + str(x))
    ###### TO CLICK I ACCEPT COOKIES ######
    try:
        browser.find_element_by_class_name("fc-button-label").click()
    except Exception as e:
        pass
    ###################################    
    time.sleep(1)
    html_source = browser.page_source  
    soup = BeautifulSoup(html_source,'html.parser')
    offers = soup.findAll('div', {'class' : 'offer_list'})


    ##### TO GET ALL OF THE LINKS WITH FLATS#######
    lol = offers[0].findAll('div', {'class': 'offer-item-in'})
    i = 0
    for lol in lol:
        links.append(lol.findAll(class_ = 'offer-link')[0].get('href'))
        i = i +1
###################################


 
for link in links:  
    table_data ={}
    count = count + 1
    i = 0
    browser.get(main_url + str(link))
    try:    
        browser.find_element_by_class_name("show-more__button d-block").click()  
        time.sleep(1)
        html_source = browser.page_source  
        soup = BeautifulSoup(html_source,'html.parser')
    except Exception as e:
        html_source = browser.page_source  
        soup = BeautifulSoup(html_source,'html.parser')
       
   
    try:
        paragraphs = []
        for x in soup:
            paragraphs.append(str(x))

        data_latitude = paragraphs[0][paragraphs[0].index('data-latitude'):paragraphs[0].index('data-longitude')].replace('data-latitude=','').replace('"', '').strip()
        table_data['latitude'] = data_latitude
   
        data_longitude = paragraphs[0][paragraphs[0].index('data-longitude'):paragraphs[0].index('data-zoom-level')].replace('data-longitude=', '').replace('"', '').strip()
        table_data['longitude'] = data_longitude
       
        nazov = soup.findAll(class_ = 'detail-title pt-4 pb-2')[0].get_text()
   
        table_data['Nazov'] = nazov
        text = soup.findAll('span', {'class': 'content-preview'})[0].get_text().replace('\n', '').replace('\xa0','')
        table_data['Text'] = text
   
        price = soup.findAll(class_ = 'contact-title big col-12 col-md-6 mb-0')[0].get_text()
        price = price[0:price.find('€')].strip()
        table_data['Cena'] = price
       
 




       
    except Exception as e:
        pass


    try:

        for x in soup.findAll('div', {'class' : 'row no-gutters mt-1'})[0]:
            key = soup.findAll('div', {'class' : 'row no-gutters mt-1'})[0].findAll('div', {'class' : 'col-sm-4 col-6 info-title'})[i].get_text()      
            value = soup.findAll('div', {'class' : 'row no-gutters mt-1'})[0].findAll('div', {'class' : 'col-sm-8 col-6'})[i].get_text()
            table_data[key] = value  
            pokus.append(key)
            pokus.append(value)
            i = i+1

    except Exception:
        pass

    i = 0
    try:
        for x in soup.findAll('div', {'class':'row no-gutters content-preview mt-1'})[0]:
            key2 = soup.findAll('div', {'class':'row no-gutters content-preview mt-1'})[0].findAll('div', {'class' : 'col-6 col-sm-4 info-title'})[i].get_text().replace('\n','').strip()
            value2 = soup.findAll('div', {'class':'row no-gutters content-preview mt-1'})[0].findAll('div', {'class' : 'col-6 col-sm-8'})[i].get_text()    
            table_data[key2] = value2
            pokus.append(key2)
            pokus.append(value2)
            i = i+1
           

    except Exception:
       
        pass

    ALL_data[count] = table_data
   

    my_dictionary = table_data

     

    filename = r'C:\Users\roboz.DESKTOP-F86F289\Desktop\picovina.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

for key in table_data.keys():
    ##f.write("%s|%s\n" % (key, my_dictionary[key]))
    ws.cell(column=1, row=ws.max_row+1, value=count)
    ws.cell(column=2, row=ws.max_row, value=key)
    ws.cell(column=3, row=ws.max_row, value=table_data[key])
    wb.save(filename)

latitude_direction = []
longitude_direction = []
cena = []
Druh = []
Typ = []
Plocha = []
nazov = []
text = []
Nazov = []
Celková_podlahová_plocha = []
Nadzemné_podlažie = []
Stav_nehnuteľnosti = []
Forma_vlastníctva = []
Konštrukcia_bytu = []
Rok_výstavby = []
Terasa = []
Energetický_certifikát_budovy = []
Internet = []
Káblová_televízia = []

for key in ALL_data:
    try:
        latitude_direction.append(ALL_data[key]['latitude'])
       
    except Exception as e:  
        latitude_direction.append('-')
   
    try:
        longitude_direction.append(ALL_data[key]['longitude'])  
    except Exception as e:
        longitude_direction.append('-')
   
    try:
        cena.append(ALL_data[key]['Cena'])
       
    except Exception as e:
        cena.append('-')
   
    try:
        Druh.append(ALL_data[key]['Druh: '])
    except Exception as e:
        Druh.append('-')
   
    try:
        Typ.append(ALL_data[key]['Typ: '])
    except Exception as e:    
        Typ.append('-')
   
    try:
        Plocha.append(ALL_data[key]['Celková úžitková plocha:'])
    except Exception as e:
        Plocha.append('-')
   
   
   
    try:
        Nadzemné_podlažie.append(ALL_data[key]['Nadzemné podlažie:'])
    except Exception as e:
        Nadzemné_podlažie.append('-')
         
    try:
        Stav_nehnuteľnosti.append(ALL_data[key]['Stav nehnuteľnosti:'])
    except Exception as e:
        Stav_nehnuteľnosti.append('-')
       
    try:
        Forma_vlastníctva.append(ALL_data[key]['Forma vlastníctva:'])
    except Exception as e:
        Forma_vlastníctva.append('-')
    try:
        Celková_podlahová_plocha.append(ALL_data[key]['Celková podlahová plocha:'])
    except Exception as e:
        Celková_podlahová_plocha.append('-')
                                                     
    try:
        Konštrukcia_bytu.append(ALL_data[key]['Konštrukcia bytu:'])
    except Exception as e:
        Konštrukcia_bytu.append('-')
       
    try:
        Rok_výstavby.append(ALL_data[key]['Rok výstavby:'])
    except Exception as e:
        Rok_výstavby.append('-')
   
    try:
        Terasa.append(ALL_data[key]['Terasa:'])
    except Exception as e:
        Terasa.append('-')
    try:
        Energetický_certifikát_budovy.append(ALL_data[key]['Energetický certifikát budovy:'])
    except Exception as e:
        Energetický_certifikát_budovy.append('-')
    try:
        Internet.append(ALL_data[key]['Internet:'])
    except Exception as e:
        Internet.append('-')
    try:
        Káblová_televízia.append(ALL_data[key]['Káblová televízia:'])
    except Exception as e:
        Káblová_televízia.append('-')
   
    try:
        Nazov.append(ALL_data[key]['Nazov'])
    except Exception as e:
        Nazov.append('-')
       
    try:
        text.append(ALL_data[key]['Text'])
    except Exception as e:
        text.append('-')    
   
                           
df = pd.DataFrame()
df['latitude_direction'] = latitude_direction
df['longitude_direction'] = longitude_direction
df['cena'] = cena
df['Druh'] = Druh
df['Typ'] = Typ
df['Nazov'] = Nazov
df['Plocha'] = Plocha
df['Celková_podlahová_plocha'] = Celková_podlahová_plocha
df['Nadzemné_podlažie'] = Nadzemné_podlažie
df['Stav_nehnuteľnosti'] = Stav_nehnuteľnosti
df['Forma_vlastníctva'] = Forma_vlastníctva
df['Konštrukcia_bytu'] = Konštrukcia_bytu
df['Rok_výstavby'] = Rok_výstavby
df['Terasa'] = Terasa
df['Energetický_certifikát_budovy'] = Energetický_certifikát_budovy
df['Internet'] = Internet
df['Káblová_televízia'] = Káblová_televízia
df['Text'] =  text  
df['Link'] = paragraphs
       

       